#Import the library of excel is openpyxl
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#We create a function where we put as a parameter the file

def process_workbook(filename):
    wb = xl.load_workbook(filename) # We called and load the file
    sheet = wb['Hoja 1'] # Called the shee on the file

    for row in range(2, sheet.max_row + 1): # Count from row 2 to max row
        cell = sheet.cell(row, 3) # We get the value on the row 3
        corrected_price = cell.value * 0.9 # The value from row 3 we multiplied 0.9
        corrected_price_cell = sheet.cell(row, 4) # The value from corrected_price we put in the row 4
        corrected_price_cell.value = corrected_price # Let's compare the value
    colors = ["blue", "purple", "green", "gray"]
    # We create a variable with the values of the sheet
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart() # We create a chart
    chart.title = "Transactions" # Title for the chart
    chart.x_axis.title = "ID Product" # Title of the part X
    chart.y_axis.title = "Price" # Title of the part Y
    chart.varyColors = colors # Color on the chart
    chart.add_data(values) # Add the values to chart
    sheet.add_chart(chart, 'f2') # We placed the chart on the F2
    wb.save(filename) # Save the file

process_workbook('transactions.xlsx') #The paramters of the function is the file
